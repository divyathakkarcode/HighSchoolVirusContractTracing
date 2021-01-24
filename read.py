from openpyxl import load_workbook
#loading excel file
wb2 = load_workbook('OEC2021-School_Record_Book.xlsx')
tas = {}
teachers = {}
studentRecord = {}

#This method will update the student, TA and teacher dictionaries with the appropriate initial infections
#No value is returned by this function as the function is simply used to update the initial infections into the "givenRisk" keys of the dictionaries
def infectionList(ListofTeachers, ListofTAs, ListofStudents):
    sheet = wb2['ZBY1 Status']
    lastRow = sheet.max_row
        
    for i in sheet["A2":"D"+str(lastRow)]:
        if(i[0].value == "N/A"):
            TA_name = i[2].value+" "+i[1].value
            for index in tas:
                if(TA_name == tas[index]["Name"]):
                    tas[index]["givenRisk"] = 1

        elif(i[0].value > 20):
            print("HERE", i[0].value)
            studentRecord[i[0].value]["givenRisk"] = 1

        else:
            teacherName = i[2].value+" "+i[1].value
            if(teacherName == teachers[i[0].value]["Name"]):
                teachers[index]["givenRisk"] = 1
            else:
                studentRecord[i[0].value]["givenRisk"] = 1

#This class creates a dictonary with all teachers
def teacherList():
    sheet = wb2['Teacher Records']
    global teachers
    lastRow = sheet.max_row
    for i in sheet["A2":"D"+str(lastRow)]:
        teachers[i[0].value] = {"TeacherNumber": i[0].value, "Name": i[2].value+" "+i[1].value, "Class": i[3].value, "givenRisk": 0.0}
        #print(teachers[i[0].value])
    return teachers

#This class creates a dictonary with all teaching assistants
def TAList():
    global tas
    sheet = wb2['Teaching Assistant Records']
    j=0
    lastRow = sheet.max_row
    for i in sheet["A2":"F"+str(lastRow)]:
        tas[j] = {"Name": i[1].value+" "+i[0].value, "Period1": i[2].value, "Period2": i[3].value, "Period3": i[4].value, "Period4": i[5].value, "givenRisk": 0.0}
        #print(tas[j])
        j+=1
    return tas

def studentList():
    sheet = wb2['Student Records']
    global studentRecord
    for i in sheet["A2":"J581"]:
        studentRecord[i[0].value] = {"StudentID": i[0].value, "Name": i[2].value+" "+i[1].value, "Grade": i[3].value, "ClassP1": i[4].value, "ClassP2": i[5].value, "ClassP3": i[6].value, "ClassP4": i[7].value, "healthFlag": 0 if(i[8].value == "N/A") else 1, "ECs": i[9].value.split(',')[0], "givenRisk": 0}
    
    return studentRecord


def putStudentsInClass():
    classesP1 = dict()
    classesP2 = dict()
    classesP3 = dict()
    classesP4 = dict()



    infectionList(teacherList(), TAList(), studentList())

    studentLst = studentRecord 
    
    for key,value in studentLst.items():
        if value['ClassP1'] in classesP1:
            classesP1[value['ClassP1']][0].append(key)
        else:
            classesP1[value['ClassP1']] = [[key]]

        if value['ClassP2'] in classesP2:
            classesP2[value['ClassP2']][0].append(key)
        else:
            classesP2[value['ClassP2']] = [[key]]

        if value['ClassP3'] in classesP3:
            classesP3[value['ClassP3']][0].append(key)
        else:
            classesP3[value['ClassP3']] = [[key]]

        if value['ClassP4'] in classesP4:
            classesP4[value['ClassP4']][0].append(key)
        else:
            classesP4[value['ClassP4']] = [[key]]


    teacherLst = teacherList()
    for key,value in teacherLst.items():
        classesP1[value['Class']].append( key)
        classesP2[value['Class']].append( key)
        classesP3[value['Class']].append( key)
        classesP4[value['Class']].append( key)


    TALst = TAList()
    for key,value in TALst.items():
        classesP1[value['Period1']].append( key)
        classesP2[value['Period2']].append( key)
        classesP3[value['Period3']].append( key)
        classesP4[value['Period4']].append( key)




    AllClasses = dict()
    AllClasses['ClassP1'] = classesP1
    AllClasses['ClassP2'] = classesP2
    AllClasses['ClassP3'] = classesP3
    AllClasses['ClassP4'] = classesP4

    return AllClasses 

print(putStudentsInClass())  
